import time
from os import makedirs

from datetime import datetime, timedelta
import math
import numpy as np
import openpyxl
import pandas as pd
from sqlalchemy import create_engine


def connect_db():
    _engine = create_engine("", encoding='utf-8')

    conn = _engine.connect()
    return conn


def load_data(sensor, sn="", stime="2022-11-21 00:00:00", avg=False):
    etime = datetime.strftime(datetime.strptime(stime, '%Y-%m-%d %H:%M:%S') + timedelta(days=1), '%Y-%m-%d %H:%M:%S')
    sql = f'''
        SELECT *
        FROM temp
        WHERE \'{stime}\' <= date AND date < \'{etime}\';
    '''

    df = pd.read_sql(sql, connect_db(), index_col=None)

    df['date'] = [d + timedelta(minutes=(5 - d.minute % 5)) if d.minute % 5 != 0 else d for d in df['date']]
    if not avg:
        df = df.drop_duplicates(['date'])
    df = df.set_index('date')

    return df.resample('5T').mean().dropna(axis=0) if avg else df


def merge_df(sensor, sn, stime, avg=False):
    df1 = load_data(sensor=sensor, sn=sn, stime=stime, avg=avg)
    df1 = df1.rename(columns={sensor: stime.split(" ")[0]})

    nStime = datetime.strftime(datetime.strptime(stime, '%Y-%m-%d %H:%M:%S') + timedelta(days=7), '%Y-%m-%d %H:%M:%S')
    df2 = load_data(sensor=sensor, sn=sn, stime=nStime, avg=avg)
    df2 = df2.rename(columns={sensor: nStime.split(" ")[0]})

    if len(df1.index)>0:
        df1.index = df1.index.strftime('%H:%M')
    if len(df2.index)>0:
        df2.index = df2.index.strftime('%H:%M')

    return pd.concat([df1, df2], axis=1)


# 엑셀파일 생성
def create_excel(extract_list, sensor_lst, stime="2022-11-21 00:00:00"):
    makedirs("result", exist_ok=True)

    for sn in extract_list:
        location = name_table.loc[name_table['sn'] == sn]
        school = location['school'].iloc[0]
        place = location['place'].iloc[0]
        makedirs(f"result/{school}", exist_ok=True)
        writer = pd.ExcelWriter(f"result/{school}/{place}_{sn}.xlsx", engine='xlsxwriter')

        for sensor in sensor_lst:
            df = merge_df(sensor=sensor, sn=sn, stime=stime)
            if len(df.index) > 290:
                df = merge_df(sensor=sensor, sn=sn, stime=stime, avg=True)
            df.to_excel(writer, sheet_name=sensor, encoding='cp949', index=True)
        writer.save()
        draw_plot(sn=sn, sensor_lst=sensor_lst, school=school, place=place)


# 생성된 엑셀파일에 그래프 그리기
def draw_plot(sn, sensor_lst, school, place):
    wb = openpyxl.load_workbook(f"result/{school}/{place}_{sn}.xlsx")

    for sensor in sensor_lst:
        ws = wb[sensor]
        chart = openpyxl.chart.LineChart()

        if ws.max_row > 1:
            data = openpyxl.chart.Reference(ws, min_col=2, min_row=1, max_col=3, max_row=ws.max_row)
            chart.add_data(data, titles_from_data=True)
            cats = openpyxl.chart.Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
            chart.set_categories(cats)

            # 그래프 사이즈 조정
            chart.height = 19.2
            chart.width = 36.2
            ws.add_chart(chart, "D1")

    wb.save(f"result/{school}/{place}_{sn}.xlsx")


if __name__ == '__main__':
    name_table = pd.read_csv("name_table.csv", encoding="utf8")

    # 추출
    extract_list = []
    # 추출해야하는 센서 목록
    sensor_list = []
    # xlsx 파일 생성
    stime = time.time()
    create_excel(sn_lst=extract_list, sensor_lst=sensor_list)
    print(time.time() - stime)
